import tensorflow as tf
import os
import keras.backend as K
import hickle as hkl
import numpy as np
import argparse
from DeepSilencer import DeepSilencer
from sklearn.utils import shuffle
from Loading_data import seq_to_kspec,checkseq,chunks,loadindex,load_genome,num2acgt,acgt2num,seq2mat,encoding_matrix
from openpyxl import load_workbook
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='DeepSilencer: Newly developed deep learning model to predict silencers')
    parser.add_argument('--data', '-d', type=str, help='input test data name',default='m_mm19_ENCODE')
    parser.add_argument('--outdir', '-o', type=str, default=os.path.dirname(os.getcwd())+'/output/crossdata-projection-mouse/', help='Output path')
    parser.add_argument('--model_name', '-f', type=str, default='../model/kmer_seq.h5', help='Model name to load for prediction')
    parser.add_argument('--mapping_file','-m',type=str,default='Mouse_mapping.xlsx',help='Mapping the cell lines we predict to their real names')
    parser.add_argument('--seed', type=int, default=1234, help='Random seed for repeat results')
    parser.add_argument('--save_result','-p', type = bool, default = True, help='Save test labels and predicted labels')
    parser.add_argument('--genome','-ge', type = str, default = 'mm10', help='The genome we need to predict')
    parser.add_argument('--start_position','-sta', type = int, default = 0, help='Start position in the data to predict')
    parser.add_argument('--end_position','-end', type = int, default = 0, help='End position in the data to predict. If set to 0, it will predict the entire test set')

    
    args = parser.parse_args()
    modelname = args.model_name
    sequences = load_genome(args.genome)
    indexes_encode = loadindex(sequences,name = args.data)
    data_name = args.data
    start_position = args.start_position
    end_position = args.end_position
    outdir = args.outdir
    mapping = args.mapping_file

    # load model
    deepsilencer = DeepSilencer()
    deepsilencer.load_weights(modelname)

    pred_result = []
    # predict the probability
    if end_position == 0 and start_position == 0:
        end_position = len(indexes_encode)
    for i in range(start_position,end_position):
        temp_sequence = indexes_encode[i]
        silencers = list()
        index = temp_sequence
        target_length = 200
        stride = 1
        # The sliding window
        try:
            [sampleid, chrkey, startpos, endpos, _] = index
        except:
            [sampleid, chrkey, startpos, endpos] = index
        origin_length = endpos - startpos
        if origin_length < target_length:
            silencer_start = startpos - target_length + origin_length
            silencer_end = endpos + target_length -origin_length
            for shift in range(0, target_length - origin_length, stride):
                start = startpos - shift
                end = start + target_length
                seq, legal = checkseq(chrkey, start, end,sequences)
                if legal:
                    silencers.append([sampleid, chrkey, start, end])
        elif origin_length >= target_length:
            silencer_start = startpos + target_length - origin_length
            silencer_end = endpos - target_length + origin_length
            chunks_ = chunks(range(startpos, endpos), target_length, target_length - stride)
            for chunk in chunks_:
                start = chunk[0]
                end = chunk[-1] + 1
                if (end - start) == target_length:
                    seq, legal = checkseq(chrkey, start, end,sequences)
                    silencers.append([sampleid, chrkey, start, end])
                elif (end - start) < target_length:
                    break
        num = len(silencers)
        silencer_mat = np.vstack([seq2mat(sequences[item[1]][item[2]:item[3]]) for item \
                                in silencers])
        silencer_mat = silencer_mat.astype(int)
        silencer_mat = silencer_mat.reshape(-1,4,200,1)
        silencer_seq = sequences[chrkey][silencer_start+1:silencer_end+1]
        test_data_kmer = []
        K = 5
        seq = silencer_seq[-201:-1]
        kmer_whole = seq_to_kspec(seq,K=K)
        kmer_whole = np.array(kmer_whole).reshape(4**K)
        kmer = np.copy(kmer_whole)
        test_data_kmer.append(kmer)
        for ind in range(num-1):
            kmer = np.copy(kmer)
            sub_seq = silencer_seq[-ind-K-1:-ind-1]
            index = 0
            for j in range(K):
                index += encoding_matrix[sub_seq[j]]*(4**(K-j-1))
            kmer[index] = kmer[index] - 1
            add_seq = silencer_seq[-ind-202:-ind-202+K]
            index = 0
            for j in range(K):
                index += encoding_matrix[add_seq[j]]*(4**(K-j-1))
            kmer[index] = kmer[index]+1
            test_data_kmer.append(kmer)
        # take the average of n results
        test_data_kmer = np.array(test_data_kmer).reshape(-1,4**K)
        pred_label = deepsilencer.predict(silencer_mat, test_data_kmer)
        pred_result.append(sum(pred_label)/num)
    
    workbook = load_workbook(mapping)    
    booksheet = workbook.active                 
    # obtain row data in sheet
    rows = booksheet.rows
    # obtain column data in sheet
    columns = booksheet.columns
    i = 1
    # Iterate over all the rows
    cell_type = []
    cell_line = []
    tissue = []
    organ = []
    for row in rows:
        i = i + 1
        line = [col.value for col in row]
        cell_data_1 = booksheet.cell(row=i, column=1).value               
        cell_data_2 = booksheet.cell(row=i, column=2).value              
        cell_data_3 = booksheet.cell(row=i, column=3).value                  
        cell_data_4 = booksheet.cell(row=i, column=4).value                   
        cell_type.append(cell_data_1)
        cell_line.append(cell_data_2)
        tissue.append(cell_data_3)
        organ.append(cell_data_4)
    type2line = {cell_type[i]:cell_line[i] for i in range(len(cell_type))}
    type2tissue = {cell_type[i]:tissue[i] for i in range(len(cell_type))}
    type2organ = {cell_type[i]:organ[i] for i in range(len(cell_type))}
    threshold = 0.5
    name_test = '%s_%d_%d_%.2f.txt'%(data_name,start_position,end_position,threshold)
    head = np.array([['Chrom','Start','End','Strand','Size','Method','Cell line','Tissue','Species'\
                    ,'Genome','Organ','Reference','Pubmed']])
    table = []
    # to predict whether it is silencer
    for i in range(start_position,end_position):
        if pred_result[i-start_position] > threshold:
            if indexes_encode[i][4] in type2line.keys():
                table.append([indexes_encode[i][1],indexes_encode[i][2]+1,indexes_encode[i][3]+1,\
                            '.',-int(indexes_encode[i][2])+int(indexes_encode[i][3]), 'DeepSilencer',type2line[indexes_encode[i][4]],\
                              type2tissue[indexes_encode[i][4]],'Mus musculus',args.genome,\
                             type2organ[indexes_encode[i][4]],'.','.'])

    table = np.array(table)
    table = np.append(head,table,axis = 0)
    # save result
    if args.save_result:
        if not os.path.exists(outdir):
            os.makedirs(outdir)
        np.savetxt(outdir + name_test,table,delimiter="\t",fmt = '%s')